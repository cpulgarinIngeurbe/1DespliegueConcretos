#!/usr/bin/env python3
"""
Script para descargar reportes de resistencia de SGS
Descarga múltiples proyectos (NATIVA, VIALE) con datos transformados
"""

from playwright.sync_api import sync_playwright
import pandas as pd
import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# Configuración
PROYECTOS = ['NATIVA', 'VIALE']
URL_LOGIN = "https://qclient.sgs.com/authentication?redirectURL=%2Fhome"
RUTA_DATOS = "datos"

# Crear carpeta datos si no existe
Path(RUTA_DATOS).mkdir(exist_ok=True)

print('[*] Iniciando descarga SGS')
print(f'[*] Proyectos: {", ".join(PROYECTOS)}')

try:
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)

        for proyecto in PROYECTOS:
            print(f'\n[*] ========== {proyecto} ==========')

            email = os.getenv(f'SGS_{proyecto}_USER')
            password = os.getenv(f'SGS_{proyecto}_PASS')

            if not email or not password:
                print(f'[!] Credenciales faltantes para {proyecto}')
                continue

            context = browser.new_context(accept_downloads=True)
            page = context.new_page()

            try:
                # Login
                page.goto(URL_LOGIN, timeout=60000)
                page.fill("input[id='user']", email)
                page.fill("input[id='password']", password)
                page.get_by_role("button", name="Ingresar").click()
                page.wait_for_selector("text=Reportes", timeout=60000)
                page.wait_for_load_state('networkidle', timeout=30000)
                print(f'[+] Login OK')

                # Navegar a Reportes
                page.locator("text=Reportes").first.click()
                page.wait_for_timeout(1500)
                page.locator("text=Generación de reportes").first.click()
                page.wait_for_timeout(1500)
                print(f'[+] En Generación de reportes')

                # Click en botón de lista
                page.locator("a.btn.btn-primary.icon-format-list-checks").click()
                page.wait_for_timeout(2000)
                page.keyboard.press("Escape")
                page.wait_for_timeout(500)

                # Seleccionar campos del formulario
                print(f'[+] Seleccionando campos...')

                # Tipo de Reporte
                page.wait_for_selector("mat-form-field:has-text('Tip de Reporte')", timeout=20000)
                page.locator("mat-form-field:has-text('Tip de Reporte') div.mat-mdc-select-trigger").click()
                page.wait_for_timeout(500)
                page.keyboard.press("ArrowDown")
                page.keyboard.press("ArrowDown")
                page.keyboard.press("ArrowDown")
                page.keyboard.press("Enter")
                page.wait_for_timeout(1000)
                print(f'[+] Tipo Reporte: OK')

                # Código de Obra
                page.locator("mat-form-field:has-text('Código de Obra') div.mat-mdc-select-trigger").click()
                page.wait_for_timeout(1000)
                page.keyboard.press("ArrowDown")
                page.keyboard.press("Enter")
                page.wait_for_timeout(1000)
                print(f'[+] Código de Obra: OK')

                # Tipo Material
                page.locator("mat-form-field:has-text('Tipo Material') div.mat-mdc-select-trigger").click()
                page.wait_for_timeout(1000)
                page.keyboard.press("ArrowDown")
                page.keyboard.press("ArrowUp")
                page.keyboard.press("Enter")
                page.wait_for_timeout(1000)
                print(f'[+] Tipo Material: OK')

                # Cerrar modales
                page.keyboard.press("Escape")
                page.keyboard.press("Escape")
                page.wait_for_timeout(800)

                # Aplicar filtro
                try:
                    page.locator("i.icon-filter-outline").click(timeout=10000)
                    page.wait_for_timeout(800)
                    print(f'[+] Filtro aplicado')
                except:
                    print(f'[!] No se pudo aplicar filtro (opcional)')

                # Marcar checkbox
                try:
                    page.check("#mat-mdc-checkbox-1-input", timeout=10000)
                    page.wait_for_timeout(500)
                    print(f'[+] Checkbox marcado')
                except:
                    print(f'[!] No se pudo marcar checkbox (opcional)')

                # Descargar Excel
                page.wait_for_selector("button:has(i.icon-file-excel)", timeout=20000)
                print(f'[+] Botón descarga encontrado')

                with page.expect_download(timeout=45000) as download_info:
                    page.locator("button:has(i.icon-file-excel)").click()

                download = download_info.value
                filepath = f'{RUTA_DATOS}/reporte_sgs_{proyecto}.xlsx'
                download.save_as(filepath)
                print(f'[+] Descargado: {filepath}')

                # Transformar datos
                print(f'[+] Transformando datos...')

                df = pd.read_excel(filepath)

                # Renombrar columnas
                mapeo = {
                    'codigoObra': 'Proyecto',
                    'asentamiento': 'Cilindro Nº',
                    'codigoMezcla': 'Código de mezcla',
                    'localizacion': 'Localización',
                    'fechaMuestreo': 'Toma',
                    'fechaRotura': 'Rotura',
                    'edad': 'Edad (días)',
                    'Resistencia Nominal Mpa': 'Resistencia nominal (MPa)',
                    'Resistencia Mpa': 'Resistencia (MPa)',
                    'porcentaje': 'Resistencia (%)'
                }

                df = df.rename(columns=mapeo)

                # Mantener columnas requeridas
                cols_mantener = [
                    'Proyecto', 'Cilindro Nº', 'Código de mezcla', 'Localización',
                    'Toma', 'Rotura', 'Edad (días)', 'Resistencia nominal (MPa)',
                    'Resistencia (MPa)', 'Resistencia (%)'
                ]
                df = df[[c for c in cols_mantener if c in df.columns]]

                # Agregar columna TIPO
                def obtener_tipo(loc):
                    if pd.isna(loc):
                        return None
                    loc_upper = str(loc).upper()
                    if "MURO" in loc_upper:
                        return "MURO"
                    elif "PLACA" in loc_upper:
                        return "PLACA"
                    elif "AUTOCOMPACTABLE" in loc_upper:
                        return "AUTOCOMPACTABLE"
                    elif "TREMIE" in loc_upper:
                        return "TREMIE"
                    return None

                df['TIPO'] = df['Localización'].apply(obtener_tipo)

                # Calcular promedios
                df['Resistencia Promedio (MPa)'] = df.groupby(
                    ['Proyecto', 'Localización', 'Cilindro Nº', 'Edad (días)']
                )['Resistencia (MPa)'].transform('mean')

                df['Conteo Elementos'] = df.groupby(
                    ['Proyecto', 'Localización', 'Cilindro Nº', 'Edad (días)']
                )['Resistencia (MPa)'].transform('count')

                df = df.rename(columns={'Resistencia (MPa)': 'Resistencia (MPa) Individual'})
                df['Resistencia (%) Individual'] = (
                    df['Resistencia (MPa) Individual'] / df['Resistencia nominal (MPa)']
                ) * 100

                # Redondear
                cols_numericas = [
                    'Edad (días)', 'Resistencia nominal (MPa)',
                    'Resistencia (MPa) Individual', 'Resistencia (%) Individual',
                    'Resistencia Promedio (MPa)', 'Resistencia (%)'
                ]
                df[cols_numericas] = df[cols_numericas].round(3)

                # Ordenar
                df = df.sort_values(
                    ['Edad (días)', 'Resistencia nominal (MPa)'],
                    ascending=[False, False]
                ).reset_index(drop=True)

                # Guardar
                df.to_excel(filepath, index=False)

                # Formatear
                wb = load_workbook(filepath)
                ws = wb.active

                azul_militar = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
                blanco = Font(color="FFFFFF", bold=True)

                for col_idx in range(1, ws.max_column + 1):
                    celda = ws.cell(1, col_idx)
                    celda.fill = azul_militar
                    celda.font = blanco

                wb.save(filepath)
                print(f'[+] ✓ Transformado y guardado')

            except Exception as e:
                print(f'[!] Error en {proyecto}: {str(e)[:100]}')
            finally:
                context.close()

        browser.close()
        print(f'\n[+] ✓ Descarga completada')

except Exception as e:
    print(f'[!] Error general: {str(e)}')
    exit(1)
