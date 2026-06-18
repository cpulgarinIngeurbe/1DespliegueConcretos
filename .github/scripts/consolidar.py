import pandas as pd
import os, json
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import subprocess

ruta_datos = "datos"
ruta_desplegable = "desplegable"
Path(ruta_desplegable).mkdir(exist_ok=True)

archivos = [f for f in os.listdir(ruta_datos) if f.endswith('.xlsx')]
print(f'[*] Consolidando {len(archivos)} archivo(s)')
dfs = []
for f in archivos:
    try:
        df = pd.read_excel(os.path.join(ruta_datos, f))
        dfs.append(df)
    except Exception as e:
        print(f'[!] Error {f}: {str(e)[:50]}')

if dfs:
    df_consolidado = pd.concat(dfs, ignore_index=True)
    ruta_consolidado = os.path.join(ruta_desplegable, 'ConsolidadoResistencias.xlsx')
    df_consolidado.to_excel(ruta_consolidado, index=False)
    wb = load_workbook(ruta_consolidado)
    ws = wb.active
    azul = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
    blanco = Font(color="FFFFFF", bold=True)
    for col in range(1, ws.max_column + 1):
        ws.cell(1, col).fill = azul
        ws.cell(1, col).font = blanco
    wb.save(ruta_consolidado)
    print(f'[+] Consolidado guardado')
    
    df = pd.read_excel(ruta_consolidado)
    proyectos = sorted(df['Proyecto'].unique().tolist()) if 'Proyecto' in df.columns else []
    df = df.astype(str)
    datos_json = {p: df[df['Proyecto'] == p].to_dict('records') for p in proyectos}
    
    options_html = '\n'.join(f'    <option value="{p}">{p}</option>' for p in proyectos)
    datos_js = json.dumps(datos_json)
    
    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Consolidado</title>
<style>
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{ font-family: Arial, sans-serif; background: #f5f5f5; color: #333; min-height: 100vh; }}
.navbar {{ background: #fff; padding: 20px 40px; border-bottom: 1px solid #ddd; }}
h1 {{ color: #4B4F1D; font-size: 24px; margin: 0; }}
.container {{ padding: 40px; max-width: 1200px; margin: 0 auto; }}
.selector {{ text-align: center; margin: 30px 0; padding: 20px; background: #fff; border-radius: 4px; }}
.selector label {{ font-weight: bold; margin-right: 15px; }}
.selector select {{ padding: 8px; font-size: 14px; border: 1px solid #ddd; }}
table {{ width: 100%; border-collapse: collapse; margin-top: 20px; background: #fff; }}
thead {{ background: #4B4F1D; color: white; }}
th {{ padding: 12px; text-align: left; font-weight: bold; }}
td {{ padding: 10px 12px; border-bottom: 1px solid #eee; }}
tbody tr:hover {{ background: #f9f9f9; }}
#contenido {{ display: none; }}
.footer {{ text-align: center; margin-top: 40px; color: #999; font-size: 12px; }}
</style>
</head>
<body>
<nav class="navbar"><h1>Consolidado de Resistencias</h1></nav>
<div class="container">
  <div class="selector">
    <label for="proyecto-select">Selecciona un proyecto:</label>
    <select id="proyecto-select">
      <option value="">-- Selecciona --</option>
{options_html}
    </select>
  </div>
  <div id="contenido">
    <table>
      <thead id="tabla-header"></thead>
      <tbody id="tabla-body"></tbody>
    </table>
  </div>
  <div class="footer">Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</div>
</div>

<script>
const datos = {datos_js};
document.getElementById('proyecto-select').addEventListener('change', function() {{
  const proyecto = this.value;
  const contenido = document.getElementById('contenido');
  if (!proyecto) {{ contenido.style.display = 'none'; return; }}
  
  const registros = datos[proyecto];
  const columns = registros.length > 0 ? Object.keys(registros[0]) : [];
  
  let headerHTML = '<tr>' + columns.map(c => '<th>' + c + '</th>').join('') + '</tr>';
  document.getElementById('tabla-header').innerHTML = headerHTML;
  
  let bodyHTML = registros.map(r => 
    '<tr>' + columns.map(c => '<td>' + (r[c] || '') + '</td>').join('') + '</tr>'
  ).join('');
  document.getElementById('tabla-body').innerHTML = bodyHTML;
  
  contenido.style.display = 'block';
}});
</script>
</body>
</html>"""
    
    with open('index.html', 'w', encoding='utf-8') as f:
        f.write(html)
    print(f'[+] Dashboard creado')
else:
    print('[!] No hay archivos')

# Hacer commit y push
try:
    token = os.getenv('GH_TOKEN', '')
    if not token:
        print('[!] GH_TOKEN no configurado')
        exit(1)
    
    subprocess.run(['git', 'config', '--global', 'user.email', 'action@github.com'], check=True)
    subprocess.run(['git', 'config', '--global', 'user.name', 'GitHub Action Bot'], check=True)
    
    subprocess.run(['git', 'add', 'desplegable/', 'index.html'], check=True)
    subprocess.run(['git', 'commit', '-m', '[AUTO] Consolidado y Dashboard'], check=False)
    
    # Obtener la URL del remoto y agregarle el token
    result = subprocess.run(['git', 'config', '--get', 'remote.origin.url'], 
                          capture_output=True, text=True, check=True)
    origin_url = result.stdout.strip()
    
    # Reemplazar con URL con token
    if 'https://' in origin_url:
        origin_url = origin_url.replace('https://', f'https://github-actions:{token}@')
    
    subprocess.run(['git', 'pull', '--rebase', 'origin', 'main'], check=True)
    subprocess.run(['git', 'push', origin_url, 'main'], check=True)
    print('[+] Push completado')
except Exception as e:
    print(f'[!] Error en push: {str(e)}')
