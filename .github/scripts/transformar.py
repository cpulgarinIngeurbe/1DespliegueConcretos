import pandas as pd
import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import subprocess

ruta_datos = "datos"
Path(ruta_datos).mkdir(exist_ok=True)

archivos_sgs = [f for f in os.listdir(ruta_datos) if f.startswith('reporte_sgs_') and f.endswith('.xlsx')]
print(f'[*] Transformando {len(archivos_sgs)} archivo(s) SGS')

for archivo in archivos_sgs:
    try:
        ruta = os.path.join(ruta_datos, archivo)
        df = pd.read_excel(ruta)
        df_renombrado = df.rename(columns={
            'codigoObra': 'Proyecto',
            'tipoMaterial': 'Material',
            'edad': 'Dias',
            'Resistencia Mpa': 'Resistencia_MPa'
        })
        df_renombrado.to_excel(ruta, index=False)
        wb = load_workbook(ruta)
        ws = wb.active
        azul = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        blanco = Font(color="FFFFFF", bold=True)
        for col in range(1, ws.max_column + 1):
            ws.cell(1, col).fill = azul
            ws.cell(1, col).font = blanco
        wb.save(ruta)
        print(f'OK: {archivo}')
    except Exception as e:
        print(f'ERROR {archivo}: {str(e)[:100]}')

archivos_concrelab = [f for f in os.listdir(ruta_datos) if f.startswith('reporte_concrelab_') and f.endswith('.xlsx')]
print(f'[*] Transformando {len(archivos_concrelab)} archivo(s) Concrelab')

for archivo in archivos_concrelab:
    try:
        ruta = os.path.join(ruta_datos, archivo)
        df = pd.read_excel(ruta, skiprows=3)
        if 'Dias de edad' in df.columns:
            df.rename(columns={'Dias de edad': 'Dias'}, inplace=True)
        if 'Resistencia (Mpa)' in df.columns:
            df.rename(columns={'Resistencia (Mpa)': 'Resistencia_MPa'}, inplace=True)
        df.to_excel(ruta, index=False)
        wb = load_workbook(ruta)
        ws = wb.active
        azul = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        blanco = Font(color="FFFFFF", bold=True)
        for col in range(1, ws.max_column + 1):
            ws.cell(1, col).fill = azul
            ws.cell(1, col).font = blanco
        wb.save(ruta)
        print(f'OK: {archivo}')
    except Exception as e:
        print(f'ERROR {archivo}: {str(e)[:100]}')

print('[+] Transformacion completada')

# Hacer commit y push
try:
    token = os.getenv('GH_TOKEN', '')
    if not token:
        print('[!] GH_TOKEN no configurado')
        exit(1)
    
    subprocess.run(['git', 'config', '--global', 'user.email', 'action@github.com'], check=True)
    subprocess.run(['git', 'config', '--global', 'user.name', 'GitHub Action Bot'], check=True)
    
    subprocess.run(['git', 'add', 'datos/'], check=True)
    subprocess.run(['git', 'commit', '-m', '[AUTO] Transformacion datos'], check=False)
    
    # Obtener la URL del remoto y agregarle el token
    result = subprocess.run(['git', 'config', '--get', 'remote.origin.url'], 
                          capture_output=True, text=True, check=True)
    origin_url = result.stdout.strip()
    
    # Reemplazar con URL con token
    if 'https://' in origin_url:
        origin_url = origin_url.replace('https://', f'https://github-actions:{token}@')
    
    subprocess.run(['git', 'pull', '--rebase', 'origin', 'main'], check=True)
    subprocess.run(['git', 'push', origin_url, 'main'], check=True)
    print('[+] Push completado')
except Exception as e:
    print(f'[!] Error en push: {str(e)}')
